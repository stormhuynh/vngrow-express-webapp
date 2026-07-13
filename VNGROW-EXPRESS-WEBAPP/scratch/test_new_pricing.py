import openpyxl
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r"C:\Users\ADMIN\.gemini\antigravity-ide\brain\45930db1-dc6c-47b1-854f-2d5dd6415652\scratch\live_sheet.xlsx", data_only=True)

# 1. Load Country Zone Mapping
country_sheet = wb['COUNTRY_ZONE_MAPPING']
countries = {}
headers_country = [country_sheet.cell(row=1, column=c).value for c in range(1, country_sheet.max_column + 1)]
for r in range(2, country_sheet.max_row + 1):
    c_name = country_sheet.cell(row=r, column=1).value
    if c_name:
        row_data = {}
        for c in range(1, country_sheet.max_column + 1):
            h = headers_country[c-1]
            if h:
                row_data[h.lower().replace(" ", "_")] = country_sheet.cell(row=r, column=c).value
        countries[c_name.upper()] = row_data

# 2. Lookup Rate Function
def lookup_rate_new(sheet_name, cargo_code, cw, zone_header):
    if sheet_name not in wb.sheetnames:
        return None
    sheet = wb[sheet_name]
    headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
    
    zone_col = -1
    zone_norm = zone_header.lower().strip()
    for idx, h in enumerate(headers):
        if h and h.lower().strip() == zone_norm:
            zone_col = idx + 1
            break
            
    if zone_col == -1:
        return None
        
    cargo_col = 1
    type_col = 2
    from_col = 3
    to_col = 4
    
    for r in range(2, sheet.max_row + 1):
        r_cargo = str(sheet.cell(row=r, column=cargo_col).value).strip().lower()
        r_type = str(sheet.cell(row=r, column=type_col).value).strip().lower()
        r_from = float(sheet.cell(row=r, column=from_col).value or 0)
        r_to = float(sheet.cell(row=r, column=to_col).value or 9999)
        rate_val = sheet.cell(row=r, column=zone_col).value
        
        if r_cargo == cargo_code.strip().lower() and r_from <= cw <= r_to:
            if rate_val and float(rate_val) > 0:
                return {
                    'rate': float(rate_val),
                    'price_type': r_type
                }
    return None

def get_group_key(s_type):
    for prefix in ['Phí kiểm dịch thực vật', 'Phí xử lý hàng xuất', 'Phụ thu Tranh tượng']:
        if prefix in s_type:
            return prefix
    return s_type

# 3. Calculate Surcharges Function
def calculate_surcharges(service, cargo_code, cw, pieces, origin, destination):
    sheet = wb['SURCHARGE']
    headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
    
    surcharges = []
    for r in range(2, sheet.max_row + 1):
        row_data = {}
        for c in range(1, sheet.max_column + 1):
            h = headers[c-1].strip().lower().replace("\t", "")
            row_data[h] = sheet.cell(row=r, column=c).value
        surcharges.append(row_data)
        
    active_surcharges = []
    norm_svc = service.upper().replace(" ", "")
    norm_cargo = cargo_code.lower().strip()
    
    for s in surcharges:
        if not s['service']:
            continue
        s_svc = str(s['service']).upper().replace(" ", "")
        if s_svc != norm_svc:
            continue
            
        s_cargo = str(s['cargo_code']).lower().strip()
        if s_cargo != 'all' and s_cargo != norm_cargo:
            continue
            
        # Check Origin in note
        note = str(s['note'] or '').lower().strip()
        if 'origin ho chi minh' in note and origin != 'HCM':
            continue
        if 'origin ha noi' in note and origin != 'HN':
            continue
            
        # Check Destination specific rules
        if s['surcharge_type'] == 'Phí hun trùng Úc' and destination.upper() != 'AUSTRALIA':
            continue
        if s['surcharge_type'] == 'Phí hun trùng các nước khác' and destination.upper() == 'AUSTRALIA':
            continue
            
        # Evaluate Condition
        cond = str(s['condition'] or '').lower().strip()
        threshold = float(s['threshold_value'] or 0)
        is_triggered = False
        multiplier = 0
        
        if cond == 'manual':
            is_triggered = True
            multiplier = 1
        elif cond == 'remote_area':
            is_triggered = False
        elif pieces:
            for p in pieces:
                l, w, h_val, gw = p.get('l', 0), p.get('w', 0), p.get('h', 0), p.get('gw', 0)
                dims = sorted([l, w, h_val], reverse=True)
                longest = dims[0]
                girth = (dims[1] + dims[2]) * 2
                lg_sum = longest + girth
                
                if cond == 'longest_side' and longest > threshold:
                    is_triggered = True
                    multiplier += 1
                elif cond == 'over_weight' and gw > threshold:
                    is_triggered = True
                    multiplier += 1
        else:
            if cond == 'piece_weight' and cw >= threshold:
                is_triggered = True
                multiplier = 1
            elif cond == 'not_over_weight' and cw < threshold:
                is_triggered = True
                multiplier = 1
                
        if is_triggered:
            amount = float(s['surcharge_amount'] or 0)
            min_amount = float(s['min_surcharge_amount'] or 0)
            basis = str(s['billing_basis'] or '').strip().lower().replace("\t", "")
            
            final_cost = 0
            if basis == 'per_kg':
                final_cost = amount * cw
            elif basis == 'per_package':
                final_cost = amount * (multiplier or 1)
            elif basis == 'per_all_package':
                final_cost = amount * cw
            elif basis == 'per_item_qty':
                qty = len(pieces) if pieces else 1
                final_cost = amount * qty
            else:
                final_cost = amount
                
            if min_amount > 0 and final_cost < min_amount:
                final_cost = min_amount
                
            active_surcharges.append({
                'surcharge_type': s['surcharge_type'],
                'amount': round(final_cost),
                'threshold_value': threshold,
                'condition': cond,
                'note': s['note']
            })
            
    # Filter tiered surcharges
    unique_active = []
    groups = {}
    for s in active_surcharges:
        gkey = get_group_key(s['surcharge_type'])
        if gkey not in groups:
            groups[gkey] = s
        else:
            existing = groups[gkey]
            # If condition is not_over_weight, keep the one with the lowest threshold
            if s['condition'] == 'not_over_weight':
                if s['threshold_value'] < existing['threshold_value']:
                    groups[gkey] = s
            else:
                # keep highest threshold (for over_weight / piece_weight)
                if s['threshold_value'] > existing['threshold_value']:
                    groups[gkey] = s
                    
    for k, val in groups.items():
        unique_active.append(val)
        
    return unique_active

# 4. Test Queries
tests = [
    {'service': 'DHL', 'cargo_code': 'normal', 'cw': 15.0, 'pieces': [], 'origin': 'HCM', 'destination': 'AFGHANISTAN'},
    {'service': 'DHL', 'cargo_code': 'cosmetics', 'cw': 15.0, 'pieces': [], 'origin': 'HCM', 'destination': 'USA'},
    {'service': 'DHL', 'cargo_code': 'cosmetics', 'cw': 8.0, 'pieces': [], 'origin': 'HN', 'destination': 'USA'},
    {'service': 'DHL', 'cargo_code': 'normal', 'cw': 85.0, 'pieces': [], 'origin': 'HCM', 'destination': 'AUSTRALIA'},
    {'service': 'DHL', 'cargo_code': 'plan', 'cw': 45.0, 'pieces': [], 'origin': 'HCM', 'destination': 'GERMANY'},
]

print("=== CALCULATOR TEST RESULTS ===")
for idx, t in enumerate(tests, 1):
    print(f"\n--- Test #{idx}: {t['service']} to {t['destination']} ({t['cw']} kg, {t['cargo_code']}, from {t['origin']}) ---")
    c_info = countries.get(t['destination'].upper())
    if not c_info:
        print("Country not found!")
        continue
    zone = c_info.get(t['service'].lower() + '_zone')
    print(f"Zone mapped: {zone}")
    res = lookup_rate_new('DHL_RATE', t['cargo_code'], t['cw'], zone)
    if not res and t['cargo_code'] != 'normal':
        print(f"Base rate for {t['cargo_code']} not found, falling back to 'normal'...")
        res = lookup_rate_new('DHL_RATE', 'normal', t['cw'], zone)
    if not res:
        print("Base rate not found!")
        continue
    base_price = res['rate']
    if res['price_type'] == 'per_kg':
        base_price *= t['cw']
    print(f"Base Freight: {base_price:,.0f} VND (Type: {res['price_type']}, Rate: {res['rate']:,.0f})")
    
    active_s = calculate_surcharges(t['service'], t['cargo_code'], t['cw'], t['pieces'], t['origin'], t['destination'])
    surch_total = 0
    print("Surcharges:")
    for s in active_s:
        print(f"  - {s['surcharge_type']}: {s['amount']:,.0f} VND ({s['note']})")
        surch_total += s['amount']
    print(f"Total Surcharges: {surch_total:,.0f} VND")
    
    subtotal = base_price + surch_total
    fuel = base_price * 0.19
    vat = (subtotal + fuel) * 0.08
    grand_total = subtotal + fuel + vat
    
    print(f"Fuel Surcharge (19% of base): {fuel:,.0f} VND")
    print(f"VAT (8% of subtotal+fuel): {vat:,.0f} VND")
    print(f"GRAND TOTAL (Sell Price): {grand_total:,.0f} VND")
